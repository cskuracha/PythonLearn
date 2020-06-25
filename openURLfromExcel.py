# Python program to create a file explorer in Tkinter 
# import all components from the tkinter library 
from tkinter import *

# import filedialog module 
from tkinter import filedialog 

import webbrowser
from openpyxl import load_workbook

# Function for opening the file explorer window 
def browseFiles(): 
    global filename
    filename = filedialog.askopenfilename(initialdir = "/", title = "Select a File", filetypes = (("XLSX files", "*.xlsx*"), ("all files", "*.*"))) 
	# Change label contents 
    #print(filename)
    label_file_explorer.configure(text="File Opened: "+filename)

def openBrowser():
    #wb = load_workbook(filename = 'E:\\Programming\\Python\\AppliedAI\\Module 3 - Chapter 2\\Sample_excel.xlsx')
    wb=load_workbook(filename=filename)
    #print(wb.sheetnames)
    sheet = wb['Sheet1']
    #print(sheet['A2'].value)
    if(start_row.get()==""):
        stRow=2
    else:
        stRow=int(start_row.get())
    if(end_row.get()==""):
        endRow=3
    else:
        endRow=int(end_row.get())
    #print(int(stRow))
    #print(int(endRow))
    # Suppose your URLs are in column 5, rows 2 to 30
    url_column = 1
    for row in range(stRow, endRow+1):
        #url = sheet.cell_value(row, url_column)
        url = sheet.cell(row, url_column).value
        webbrowser.open_new_tab(url)
	
# Create the root window 
window = Tk() 


# Set window title 
window.title('File Explorer') 

# Set window size 
window.geometry("600x600") 

#Set window background color 
window.config(background = "white") 

#frame=Frame(window)
#frame.pack()


# Create a File Explorer label 
label_file_explorer = Label(window, text = "File Explorer", width = 100, height = 4, fg = "blue") 
#label_file_explorer.pack()	
label_file_explorer.place(x=10,y=10)

label_explorer=Label(window,text="Browse File Name")
#label_explorer.pack(side=LEFT)
label_explorer.place(x=10,y=100)

button_explore = Button(window, text = "Browse Files", command = browseFiles) 
#button_explore.pack(side=RIGHT)
button_explore.place(x=150,y=100)

start_row=Entry(window,bd=5)
#start_row.pack(side=LEFT)
start_row.place(x=10,y=170)

end_row=Entry(window,bd=5)
#end_row.pack(side=RIGHT)
end_row.place(x=200,y=170)

button_open_browser=Button(window,text="Open Browser",command=openBrowser)
#button_open_browser.pack(side=BOTTOM)
button_open_browser.place(x=400,y=170)

button_exit = Button(window, text = "Exit", command = exit) 
#button_exit.pack(side=BOTTOM)
button_exit.place(x=10,y=230)
"""
# Grid method is chosen for placing the widgets at respective positions in a table like structure by specifying rows and columns 
label_file_explorer.grid(column = 1, row = 1,ipady=2) 

label_explorer.grid(column=1,row=2,ipady=2)
button_explore.grid(column = 2, row = 2,ipady=2) 
start_row.grid(column=1,row=3,ipady=2)
end_row.grid(column=2,row=3,ipady=2)
button_open_browser.grid(column=3, row=3,ipady=2)
button_exit.grid(column = 1,row = 4,ipady=2) 
"""
# Let the window wait for any events 
window.mainloop()
